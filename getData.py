import mysql.connector
from openpyxl import load_workbook
import tqdm
from tqdm import tqdm



conn = mysql.connector.connect(
    host = "host",
    user = "user",
    password = "password",
    database = "database-name",
)
cursor = conn.cursor()

for n in tqdm(range(2,10)):
    file_path = "food{}.xlsx".format(n)

    wb = load_workbook(file_path)
    ws = wb.active

    name = None
    energy = None
    protein = None
    fat = None
    carbohydrate = None
    sugar = None
    sodium = None
    cholesterol = None
    saturate_fat = None
    trans_fat = None
    brand_id = None
    category_id = None
    hydrate  = None
    saccharose  = None
    glucose  = None
    fructose  = None
    lactose  = None
    maltose  = None
    dietary_fiber  = None
    calcium  = None
    iron  = None
    magnesium  = None
    phosphorus  = None
    kalium  = None
    zinc  = None
    copper  = None
    manganese  = None
    selenium = None
    retinol = None
    beta_carotene  = None
    vitamin_D3  = None
    tocopherol  = None
    tocotrienols  = None
    vitamin_B1  = None
    vitamin_B2  = None
    niacin  = None
    folate  = None
    vitamin_B12  = None
    vitamin_C  = None
    amino_acid  = None
    isoleucine  = None
    leucine  = None
    lysine  = None
    methionine  = None
    phenylalanine  = None
    threonine  = None
    valine  = None
    histidine  = None
    arginine  = None
    tyrosine  = None
    cysteine  = None
    alanine  = None
    aspartic_acid  = None
    glutamic_acid  = None
    glycine  = None
    proline  = None
    serine  = None
    butyric_acid  = None
    caproic_acid  = None
    caprylic_acid  = None
    capric_acid  = None
    lauric_acid  = None
    myristic_acid  = None
    palmitic_acid  = None
    stearic_acid  = None
    arachidic_acid  = None
    myristoleic_acid  = None
    palmitoleic_acid  = None
    oleic_acid  = None
    vaccenic_acid  = None
    gadoleic_acid  = None
    linoleic_acid  = None
    alpha_linolenic_acid  = None
    gamma_linolenic_acid  = None
    eicosadienoic_acid  = None
    arachidonic_acid  = None
    eicosatrienoic_acid  = None
    eicosapentaenoic_acid  = None
    docosapentaenoic_acid  = None
    docosahexaenoic_acid  = None
    trans_oleic_acid  = None
    trans_linoleic_acid  = None
    trans_linolenic_acid  = None
    ash  = None
    caffeine = None
    sugar_alcohol = None
    erythritol = None
    iodine = None
    chloride = None
    vitamin_D = None
    vitamin_D1 = None
    vitamin_E = None
    vitamin_K  = None
    vitamin_K1  = None
    vitamin_K2  = None
    pantothenic_acid  = None
    vitamin_B6  = None
    biotin  = None
    choline  = None
    tryptophan  = None
    taurine = None
    omega_3_fatty_acids = None
    total_unsaturated_fats = None

    portion_size_idx = None
    name_idx = None
    energy_idx = None
    protein_idx = None
    fat_idx = None
    carbohydrate_idx = None
    sugar_idx = None
    sodium_idx = None
    cholesterol_idx = None
    saturate_fat_idx = None
    trans_fat_idx = None
    category_idx = None
    brand_idx = None
    hydrate_idx = None
    saccharose_idx = None
    glucose_idx = None
    fructose_idx = None
    lactose_idx = None
    maltose_idx = None
    dietary_fiber_idx = None
    calcium_idx = None
    iron_idx = None
    magnesium_idx = None
    phosphorus_idx = None
    kalium_idx = None
    zinc_idx = None
    copper_idx = None
    manganese_idx = None
    selenium_idx = None
    retinol_idx = None
    beta_carotene_idx = None
    vitamin_D3_idx = None
    tocopherol_idx = None
    tocotrienols_idx = None
    vitamin_B1_idx = None
    vitamin_B2_idx = None
    niacin_idx = None
    folate_idx = None
    vitamin_B12_idx = None
    vitamin_C_idx = None
    amino_acid_idx = None
    isoleucine_idx = None
    leucine_idx = None
    lysine_idx = None
    methionine_idx = None
    phenylalanine_idx = None
    threonine_idx = None
    valine_idx = None
    histidine_idx = None
    arginine_idx = None
    tyrosine_idx = None
    cysteine_idx = None
    alanine_idx = None
    aspartic_acid_idx = None
    glutamic_acid_idx = None
    glycine_idx = None
    proline_idx = None
    serine_idx = None
    butyric_acid_idx = None
    caproic_acid_idx = None
    caprylic_acid_idx = None
    capric_acid_idx = None
    lauric_acid_idx = None
    myristic_acid_idx = None
    palmitic_acid_idx = None
    stearic_acid_idx = None
    arachidic_acid_idx = None
    myristoleic_acid_idx = None
    palmitoleic_acid_idx = None
    oleic_acid_idx = None
    vaccenic_acid_idx = None
    gadoleic_acid_idx = None
    linoleic_acid_idx = None
    alpha_linolenic_acid_idx = None
    gamma_linolenic_acid_idx = None
    eicosadienoic_acid_idx = None
    arachidonic_acid_idx = None
    eicosatrienoic_acid_idx = None
    eicosapentaenoic_acid_idx = None
    docosapentaenoic_acid_idx = None
    docosahexaenoic_acid_idx = None
    trans_oleic_acid_idx = None
    trans_linoleic_acid_idx = None
    trans_linolenic_acid_idx = None
    ash_idx = None
    caffeine_idx = None
    sugar_alcohol_idx = None
    erythritol_idx = None
    iodine_idx = None
    chloride_idx = None
    vitamin_D_idx = None
    vitamin_D1_idx = None
    vitamin_E_alpha_TE_idx = None
    vitamin_K_mg_idx = None
    vitamin_K1_idx = None
    vitamin_K2_idx = None
    pantothenic_acid_idx = None
    vitamin_B6_idx = None
    biotin_idx = None
    choline_idx = None
    tryptophan_idx = None
    taurine_idx = None
    omega_3_fatty_acids_idx = None
    total_unsaturated_fats_idx = None

    food_data = [category_id, brand_id, name, energy, protein, fat, carbohydrate, sugar, sodium, cholesterol, saturate_fat, trans_fat,
                    hydrate , saccharose , glucose , fructose , lactose , maltose , 
                    dietary_fiber , calcium , iron , magnesium , phosphorus , kalium , 
                    zinc , copper , manganese , selenium ,retinol, beta_carotene , vitamin_D3 , 
                    tocopherol , tocotrienols , vitamin_B1 , vitamin_B2 , niacin , folate , 
                    vitamin_B12 , vitamin_C , amino_acid , isoleucine , leucine ,  
                    lysine , methionine , phenylalanine , threonine , valine , histidine , 
                    arginine , tyrosine ,  cysteine , alanine , aspartic_acid , 
                    glutamic_acid , glycine , proline , serine , butyric_acid , 
                    caproic_acid , caprylic_acid , capric_acid , lauric_acid , myristic_acid , 
                    palmitic_acid , stearic_acid , arachidic_acid , myristoleic_acid , 
                    palmitoleic_acid , oleic_acid , vaccenic_acid , gadoleic_acid , 
                    linoleic_acid , alpha_linolenic_acid , gamma_linolenic_acid , eicosadienoic_acid , 
                    arachidonic_acid , eicosatrienoic_acid , eicosapentaenoic_acid , 
                    docosapentaenoic_acid , docosahexaenoic_acid , trans_oleic_acid , 
                    trans_linoleic_acid , trans_linolenic_acid , ash , caffeine, sugar_alcohol, erythritol, 
                    iodine, chloride, vitamin_D, vitamin_D1,vitamin_E , vitamin_K , 
                    vitamin_K1 , vitamin_K2 , pantothenic_acid , vitamin_B6 , biotin , choline , tryptophan , 
                    taurine, omega_3_fatty_acids, total_unsaturated_fats]

    food_idx = [portion_size_idx, category_idx, brand_idx, name_idx, energy_idx, protein_idx, fat_idx, carbohydrate_idx, sugar_idx, 
                sodium_idx, cholesterol_idx, saturate_fat_idx, trans_fat_idx,
                hydrate_idx, saccharose_idx, glucose_idx, fructose_idx, lactose_idx, maltose_idx, dietary_fiber_idx, 
                calcium_idx, iron_idx, magnesium_idx, phosphorus_idx, kalium_idx, zinc_idx, copper_idx, manganese_idx, 
                selenium_idx, retinol_idx, beta_carotene_idx, vitamin_D3_idx, tocopherol_idx, tocotrienols_idx, 
                vitamin_B1_idx, vitamin_B2_idx, niacin_idx, folate_idx, vitamin_B12_idx, vitamin_C_idx, amino_acid_idx, 
                isoleucine_idx, leucine_idx, lysine_idx, methionine_idx, phenylalanine_idx, threonine_idx, valine_idx, 
                histidine_idx, arginine_idx, tyrosine_idx, cysteine_idx, alanine_idx, aspartic_acid_idx, glutamic_acid_idx, 
                glycine_idx, proline_idx, serine_idx, butyric_acid_idx, caproic_acid_idx, caprylic_acid_idx, capric_acid_idx, 
                lauric_acid_idx, myristic_acid_idx, palmitic_acid_idx, stearic_acid_idx, arachidic_acid_idx, myristoleic_acid_idx, 
                palmitoleic_acid_idx, oleic_acid_idx, vaccenic_acid_idx, gadoleic_acid_idx, linoleic_acid_idx, 
                alpha_linolenic_acid_idx, gamma_linolenic_acid_idx, eicosadienoic_acid_idx, arachidonic_acid_idx, 
                eicosatrienoic_acid_idx, eicosapentaenoic_acid_idx, docosapentaenoic_acid_idx, docosahexaenoic_acid_idx, 
                trans_oleic_acid_idx, trans_linoleic_acid_idx, trans_linolenic_acid_idx, ash_idx, caffeine_idx, sugar_alcohol_idx, 
                erythritol_idx, iodine_idx, chloride_idx, vitamin_D_idx, vitamin_D1_idx, 
                vitamin_E_alpha_TE_idx, vitamin_K_mg_idx, vitamin_K1_idx, vitamin_K2_idx, pantothenic_acid_idx, 
                vitamin_B6_idx, biotin_idx, choline_idx, tryptophan_idx, taurine_idx, omega_3_fatty_acids_idx, total_unsaturated_fats_idx]

    column_names =  ('1회제공량', '식품상세분류', '제조사', '식품명', '에너지(㎉)', '단백질(g)', '지방(g)', '탄수화물(g)', '총당류(g)', '나트륨(㎎)', '콜레스테롤(㎎)', 
            '총 포화 지방산(g)', '트랜스 지방산(g)', '수분(g)', '자당(g)', '포도당(g)', '과당(g)', '유당(g)', 
            '맥아당(g)', '총 식이섬유(g)', '칼슘(㎎)', '철(㎍)', '마그네슘(㎎)', '인(㎎)', '칼륨(㎎)', '아연(㎎)', '구리(㎎)', 
            '망간(㎎)', '셀레늄(㎍)', '비타민 A(㎍ RE)', '베타카로틴(㎍)', '비타민 D3(㎍)', '토코페롤(㎎)', '토코트리에놀(㎎)', '비타민 B1(㎎)', 
            '비타민 B2(㎎)', '나이아신(㎎)', '엽산(DFE)(㎍)', '비타민 B12(㎍)', '비타민 C(㎎)', '총 아미노산(㎎)', '이소류신(㎎)', '류신(㎎)', 
            '라이신(㎎)', '메티오닌(㎎)', '페닐알라닌(㎎)', '트레오닌(㎎)', '발린(㎎)', '히스티딘(㎎)', '아르기닌(㎎)', '티로신(㎎)', '시스테인(㎎)', 
            '알라닌(㎎)', "아스파르트산(㎎)", "글루탐산(㎎)", "글리신(㎎)", "프롤린(㎎)", "세린(㎎)", "부티르산(4:0)(g)", "카프로산(6:0)(g)", 
            "카프릴산(8:0)(g)", "카프르산(10:0)(g)", "라우르산(12:0)(g)", "미리스트산(14:0)(g)", "팔미트산(16:0)(g)", "스테아르산(18:0)(g)", 
            "아라키드산(20:0)(g)", "미리스톨레산(14:1)(g)", "팔미톨레산(16:1)(g)", "올레산(18:1(n-9))(g)", "박센산(18:1(n-7))(g)", "가돌레산(20:1)(g)", 
            "리놀레산(18:2(n-6)c)(g)", "알파 리놀렌산(18:3(n-3))(㎎)", "감마 리놀렌산(18:3(n-6))(㎎)", "에이코사디에노산(20:2(n-6))(g)", "아라키돈산(20:4(n-6))(㎎)", 
            "에이코사트리에노산(20:3(n-6))(g)", "에이코사펜타에노산(20:5(n-3))(㎎)", "도코사펜타에노산(22:5(n-3))(g)", "도코사헥사에노산(22:6(n-3))(㎎)", 
            "트랜스 지방산(g)", "트랜스 올레산(18:1(n-9)t)(g)", "트랜스 리놀레산(18:2t)(g)", "회분(g)", "카페인(㎎)", '당알콜(g)', '에리스리톨(g)', '요오드(㎍)', 
            '염소(㎎)', '비타민 D(D2+D3)(㎍)', '비타민 D1(㎍)', '비타민 E(㎎ α-TE)', '비타민 K(㎎)', 
            '비타민 K1(㎍)', '비타민 K2(㎍)', '판토텐산(㎎)', '비타민 B6(㎎)', '비오틴(㎍)', '콜린(㎎)', '트립토판(㎎)', '타우린(㎎)', '오메가 3 지방산(g)', 
            '총 불포화지방산(g)',)

    def is_number(s):
        try:
            float(s)
            return True
        except ValueError:
            return False
        
    columns = list(ws.iter_rows(min_row=1, max_row=1,values_only=True))[0]
    for idx, column_name in enumerate(columns):
        for i ,col_n in enumerate(column_names):
            if col_n == '제조사':
                if col_n in column_name:
                    food_idx[i] = idx
            elif column_name == col_n:
                food_idx[i] = idx

    for row in tqdm(ws.iter_rows(min_row=2, values_only=True), total=ws.max_row - 1):
        category_query = "INSERT IGNORE INTO Category (name) VALUES (%s)"
        category_data = (row[food_idx[1]],)
        brand_query = "INSERT IGNORE Brand (name) VALUES (%s)"
        brand_data = (row[food_idx[2]],)
        cursor.execute(category_query, category_data)
        cursor.execute(brand_query, brand_data)
        cursor.execute("SELECT id FROM Category WHERE name = %s", (row[food_idx[1]],))
        food_data[0] = cursor.fetchone()[0]
        cursor.execute("SELECT id FROM Brand WHERE name = %s", (row[food_idx[2]],))
        food_data[1] = cursor.fetchone()[0]

        div = float(row[food_idx[0]]) / 100
        if row[food_idx[3]] == '-':
            food_data[2] = None
        else: 
            food_data[2] = row[food_idx[3]]
        for index, nut_idx in enumerate(food_idx[4:], start=3):
            if nut_idx is not None:
                food_data[index] = row[nut_idx]
                if is_number(food_data[index]) == False:
                    food_data[index] = None
                else:
                    food_data[index] = float(row[nut_idx]) / div

        food_query = """INSERT INTO Food (category_id, brand_id, name, energy, protein, fat, carbohydrate, sugar, sodium, 
                    cholesterol, saturate_fat, trans_fat,  
                    hydrate , saccharose , glucose , fructose , lactose , maltose , 
                    dietary_fiber , calcium , iron , magnesium , phosphorus , kalium , 
                    zinc , copper , manganese , selenium ,retinol, beta_carotene , vitamin_D3 , 
                    tocopherol , tocotrienols , vitamin_B1 , vitamin_B2 , niacin , folate , 
                    vitamin_B12 , vitamin_C , amino_acid , isoleucine , leucine ,  
                    lysine , methionine , phenylalanine , threonine , valine , histidine , 
                    arginine , tyrosine ,  cysteine , alanine , aspartic_acid , 
                    glutamic_acid , glycine , proline , serine , butyric_acid , 
                    caproic_acid , caprylic_acid , capric_acid , lauric_acid , myristic_acid , 
                    palmitic_acid , stearic_acid , arachidic_acid , myristoleic_acid , 
                    palmitoleic_acid , oleic_acid , vaccenic_acid , gadoleic_acid , 
                    linoleic_acid , alpha_linolenic_acid , gamma_linolenic_acid , eicosadienoic_acid , 
                    arachidonic_acid , eicosatrienoic_acid , eicosapentaenoic_acid , 
                    docosapentaenoic_acid , docosahexaenoic_acid , trans_oleic_acid , 
                    trans_linoleic_acid , trans_linolenic_acid , ash , caffeine, sugar_alcohol, erythritol, 
                    iodine, chloride, vitamin_D, vitamin_D1,vitamin_E , vitamin_K , 
                    vitamin_K1 , vitamin_K2 , pantothenic_acid , vitamin_B6 , biotin , choline , tryptophan , 
                    taurine, omega_3_fatty_acids, total_unsaturated_fats)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
                            %s, %s, %s, %s, %s, %s, %s, %s)"""
        cursor.execute(food_query, food_data)

conn.commit()
cursor.close()
conn.close()

